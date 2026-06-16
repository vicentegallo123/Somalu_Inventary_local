"""
exportar.py — Genera reporte Excel desde la base de datos SQLite
"""
import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from rich.console import Console
from rich.prompt import Prompt
from modules import db

console = Console()
EXPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "exports")


def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def _estilo_header(ws, fila=1, color="1F3864"):
    for cell in ws[fila]:
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = PatternFill("solid", start_color=color)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def exportar_excel():
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    fecha = datetime.now().strftime("%Y%m%d_%H%M")
    nombre_default = f"Inventario_SOMALU_{fecha}.xlsx"
    nombre = Prompt.ask("Nombre del archivo", default=nombre_default)
    if not nombre.endswith(".xlsx"):
        nombre += ".xlsx"
    ruta = os.path.join(EXPORTS_DIR, nombre)

    # ── Hoja 1: Inventario ────────────────────────────────
    rows_prod = db.get_todos_productos()
    df_prod = pd.DataFrame(rows_prod, columns=["ID", "Categoría", "Producto", "Stock", "Unidad", "Costo/U"])
    df_prod["Total Costo"] = df_prod["Stock"] * df_prod["Costo/U"]
    df_prod = df_prod.drop(columns=["ID"])

    # ── Hoja 2: Movimientos ───────────────────────────────
    rows_mov = db.get_movimientos(limite=1000)
    df_mov = pd.DataFrame(rows_mov, columns=["Fecha", "Tipo", "Folio", "Proveedor/Motivo", "Producto", "Cantidad", "Costo/U", "Notas"])

    # ── Hoja 3: Proveedores ───────────────────────────────
    rows_prov = db.get_proveedores()
    df_prov = pd.DataFrame(rows_prov, columns=["ID", "Nombre", "Teléfono", "Email", "Comentarios"])
    df_prov = df_prov.drop(columns=["ID"])

    with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
        df_prod.to_excel(writer, sheet_name="Inventario", index=False)
        df_mov.to_excel(writer, sheet_name="Movimientos", index=False)
        df_prov.to_excel(writer, sheet_name="Proveedores", index=False)

    # ── Formato ───────────────────────────────────────────
    wb = load_workbook(ruta)

    for nombre_hoja, color in [("Inventario", "1F3864"), ("Movimientos", "7B3F00"), ("Proveedores", "1A5276")]:
        ws = wb[nombre_hoja]
        _estilo_header(ws, 1, color)
        _autofit(ws)
        ws.freeze_panes = "A2"

    # Totales en Inventario
    ws_inv = wb["Inventario"]
    last = ws_inv.max_row + 1
    ws_inv[f"D{last}"] = f"=SUM(D2:D{last-1})"
    ws_inv[f"G{last}"] = f"=SUM(G2:G{last-1})"
    for c in [f"D{last}", f"G{last}"]:
        ws_inv[c].font = Font(bold=True, name="Arial")

    wb.save(ruta)
    console.print(f"[bold blue]🖨️  Reporte guardado en:[/bold blue] [green]{ruta}[/green]")
